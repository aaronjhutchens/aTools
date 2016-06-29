#!/usr/bin/python

import sys, time, openpyxl
from openpyxl import Workbook

newFile = sys.argv[2]
oldFile = sys.argv[1]

print "Loading old file: ", oldFile
oldWs = openpyxl.load_workbook(oldFile).active
print "Loading new file: ", newFile
newWs = openpyxl.load_workbook(newFile).active
found = False

destFile = "postComparison_" + time.strftime("%m_%d_%y") + ".xlsx"
wb = Workbook()

totalNewRows = newWs.max_row
totalNewCols = newWs.max_column

totalOldRows = oldWs.max_row
totalOldCols = oldWs.max_column

#for col in range(1, totalOldCols + 1):
#    print oldWs.cell(row=1, column=(col)).value


def delta(oldValue, newValue):
    "Simply returns a difference"
    try:
        value = float(newValue - oldValue)
        return value
    except TypeError:
        value = "NaN"
        return value

def deltaPercent(oldValue, newValue):
    "Returns percent increase/decrease"
    if oldValue == 0:
        print "Divide by zero!"
        value = "div/0"
    else:
        try:
            value = float(newValue / oldValue * 100)
            return value
        except TypeError:
            value = "NaN"
            return value


for newRow in range(1, totalNewRows + 1):
    # This is ugly and shitty and I hate this
    department  = newWs.cell(row=(newRow), column=2).value
    office      = newWs.cell(row=(newRow), column=3).value
    newName     = newWs.cell(row=(newRow), column=1).value
    if department == "Unix Management" and office == "Chicago":
        for oldRow in range(1, totalOldRows + 1):
            oldName = oldWs.cell(row=(oldRow), column=1).value
            if newName == oldName:
                ws = wb.create_sheet(title=newName) 
                found = True
                ws.cell(row=1, column=1).value  = newName
                ws.cell(row=1, column=2).value  = "Old data"
                ws.cell(row=1, column=3).value  = "Delta"
                ws.cell(row=1, column=4).value  = "Delta %"
                ws.cell(row=1, column=5).value  = "New Data"
                for col in range(2, totalOldCols):
                    ws.cell(row=col, column=1).value = oldWs.cell(row=1, column=(col)).value
                    ws.cell(row=col, column=2).value = oldWs.cell(row=oldRow, column=(col)).value
                    ws.cell(row=(col - 1), column=5).value = newWs.cell(row=newRow, column=(col - 1)).value

                    if col > 3:
                        print col
                        ws.cell(row=col, column=3).value = delta(oldWs.cell(row=oldRow, column=(col)).value, newWs.cell(row=oldRow, column=(col)).value)
                        ws.cell(row=col, column=4).value = deltaPercent(oldWs.cell(row=oldRow, column=(col)).value, newWs.cell(row=oldRow, column=(col)).value)


            if found == False:
                print "No matching values found for ", newName, oldName

wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
wb.save(filename = destFile)
print "File saved as: ", destFile


